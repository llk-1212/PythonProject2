from selenium import webdriver
import time
from openpyxl import Workbook
import openpyxl
import json
import requests
from bs4 import BeautifulSoup
import bs4
from random import randint
import random
import re
def excel():
    excelpath = r'E:\xl.xlsx'
    wtbook = Workbook()
    # 新增一个sheet工作表
    sheet = wtbook.active
    # 写入数据头
    row = 1
    col = 1
    headlist = [u'项目名称', u'项目类型', u'已筹金额', u'捐款人次', u'发布时间',
                u'项目详情', u'发起方', u'善款接收', u'募捐动态', u'进展报告']
    for head in headlist:
        sheet.cell(row, col, head)
        col = col + 1
    wtbook.save(excelpath)

def fes():
    lis = ips()
    random_ip = lis[randint(0, len(lis) - 1)]
    requests.adapters.DEFAULT_RETRIES = 5
    s = requests.session()
    s.keep_alive = False
    USER_AGENTS = [
        "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; AcooBrowser; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
        "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; Acoo Browser; SLCC1; .NET CLR 2.0.50727; Media Center PC 5.0; .NET CLR 3.0.04506)",
        "Mozilla/4.0 (compatible; MSIE 7.0; AOL 9.5; AOLBuild 4337.35; Windows NT 5.1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
        "Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)",
        "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
        "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
        "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
        "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
        "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5",
        "Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.8) Gecko Fedora/1.9.0.8-1.fc10 Kazehakase/0.5.6",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
        "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52",
    ]
    random_agent = USER_AGENTS[randint(0, len(USER_AGENTS) - 1)]
    head = {
        'User-Agent': random_agent,
    }
    excelpath = r'E:\xl.xlsx'
    wtbook = openpyxl.load_workbook(excelpath)
    sheet = wtbook.worksheets[0]
    n = 0
    for i in range(182,200):
        ur = "https://gongyi.weibo.com/list/personal?on_state=0&donate_type=0&state=1&type=0&location=&title=&open=0" \
             "&page=" + str(i)
        s.headers = head
        s.proxies = {"http": random_ip}
        res = s.get(ur)
        soup = BeautifulSoup(res.text,'html.parser')
        lis = soup.select("div[class='person_project clearfix']")
        for li in lis:
            n = n + 1
            print(n)
            row = sheet.max_row
            row = row + 1
            div = li.find(name='div',attrs={'class:','view_info'})
            name = div.find(name='div',attrs={'class:','title'})
            title = name.find('a')
            id = title.get('href').split('/')[1]
            url = "https://gongyi.weibo.com/" + id
            response = s.get(url,headers=head, verify=False)
            soups = BeautifulSoup(response.text,'html.parser')
            titles = soups.find(name='h3',attrs={'class:', 'tit'}).find('strong')
            sheet.cell(row, 1, titles.string)
            lei = soups.find(name='span',attrs={'class:','tagging_a'})
            sheet.cell(row,2,lei.string)
            ia = soups.findAll(name='i',attrs={'class:','m-fontWeight'})
            sheet.cell(row, 3, ia[0].text)
            sheet.cell(row, 4, ia[1].text)
            shi = soups.findAll(name='div',attrs={'class:','txt'})
            shis = shi[1].text.split('-')
            sheet.cell(row, 5, shis[0])
            nei = soups.find(name='div',attrs={'class:','tab_a_con'})
            nei = tosting(nei)
            sheet.cell(row, 6, nei.text.split(sep='展开全文')[0].replace("\n",' '))
            fa = soups.find(name='div',attrs={'class:','b_con_conL'})
            fa = tosting(fa)
            sheet.cell(row, 7, fa.text.replace("\n",' '))
            shan =  soups.find(name='div',attrs={'class:','b_con_conR'})
            shan = tosting(shan)
            a = shan.find(name='a',id='show_persion')
            b = shan.find(name='a',id='show_public')
            a.string = a.get('url_open')
            b.string = b.get('url_open')
            sheet.cell(row, 8, shan.text.replace("\n", ' '))
            jia = soups.find(name='div',attrs={'class:','con_loading'})
            stings = ''
            if jia is not None:
                muu = "https://gongyi.weibo.com/aj_personal_getdonatelist?pageSize=10&page=1" + "&help_id=" + id
                reso = s.get(muu, headers={'referer': url, 'User-Agent': random_agent}, verify=False)
                js = json.loads(reso.text)
                num = (int(js['data']['total']) / 10) + 1
                for x in range(1,int(num)):
                    muurl = "https://gongyi.weibo.com/aj_personal_getdonatelist?pageSize=10&page=" \
                          + str(x) + "&help_id=" + id
                    resp = s.get(muurl,headers={'referer': url, 'User-Agent': random_agent}, verify=False)
                    jso = json.loads(resp.text)
                    lise = jso['data']['donate_list']
                    for k in range(0,len(lise)):
                        lises = lise[k]
                        stings = stings + lises['profile_image_url'] + ' ' + lises['screen_name'] + ' ' + lises['pay_time'] + ' ' + lises['msg'] + ' ' + lises['money'] + ' '
            else:
                mu = soups.find(name='div',attrs={'class:','tab_dynamic'})
                if mu is not None:
                    mu = tosting(mu)
                    stings = stings + mu.text.replace("\n", ' ')
            sheet.cell(row, 9, stings)
            jin = soups.find(name='div',attrs={'class:','tab_progress'})
            if jin is not None:
                jin = tosting(jin)
                sheet.cell(row, 10, jin.text.replace("\n", ' ').replace("确认删除？ 删除取消", ' '))
            else:
                sheet.cell(row, 10, '')
        wtbook.save(excelpath)


def tosting(str):
    src = str.findAll(name='img')
    for i in range(0, len(src)):
        src[i].string = src[i].get('src') + ' '
    return str


def ips():
    list = []
    s = requests.session()
    s.keep_alive = False
    for i in range(2, 40):
        url = "http://www.66ip.cn/" + str(i) + ".html"
        r = s.get(url)
        soup = BeautifulSoup(r.text, 'html.parser')
        tbody = soup.find(name='body')
        trs = tbody.findAll(name='tr')
        for i in range(2, len(trs)):
            tr = trs[i]
            tds = tr.findAll(name='td')
            ur = "http://" + tds[0].text + ":" + tds[1].text
            li = "https://123.sogou.com/"
            try:
                rs = s.get(li, proxies={"http": ur})
                list.append(ur)
            except:
                print("该代理不可用")
    return list


if __name__ == '__main__':
    fes()
