import json
import re

import openpyxl
import requests
from openpyxl import Workbook
from random import randint



def excel():
    wtbook = Workbook()
    # 新增一个sheet工作表
    sheet = wtbook.active
    # 写入数据头
    row = 1
    col = 1
    headlist = [u'项目名称', u'项目类型（综合/济困救灾/疾病救助)', u'项目状态（募款中/执行中/已结束)', u'项目实施时间段', u'项目开始时间', u'捐款人次', u'已筹金额', u'目标金额',
                u'筹款百分比', u'公益机构', u'发起方', u'执行方', u'公募支持', u'项目介绍', u'项目预算', u'执行计划', u'项目效果', u'执行能力说明', u'关于我们']
    for head in headlist:
        sheet.cell(row, col, head)
        col = col + 1
    wtbook.save(r'E:\le.xlsx')


def pri():
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    USER_AGENTS = [
        "Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.87 Safari/537.36 SE 2.X MetaSr 1.0",
        "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; Acoo Browser; SLCC1; .NET CLR 2.0.50727; Media Center PC 5.0; .NET CLR 3.0.04506)",
        "Mozilla/4.0 (compatible; MSIE 7.0; AOL 9.5; AOLBuild 4337.35; Windows NT 5.1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
        "Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)",
        "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
        "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
        "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
        "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
        "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5",
        "Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.8) Gecko Fedora/1.9.0.8-1.fc10 Kazehakase/0.5.6",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
        "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52",
    ]
    random_agent = USER_AGENTS[randint(0, len(USER_AGENTS) - 1)]
    head = {
        'User-Agent': random_agent,
    }
    excelpath = r'E:\le.xlsx'
    wtbook = openpyxl.load_workbook(excelpath)
    sheet = wtbook.worksheets[0]
    for j in range(3,4):
        for m in range(3, 6):
            for i in range(1, 101):
                urls = "https://ssl.gongyi.qq.com/cgi-bin/WXSearchCGI?ptype=stat&s_status=" + str(j) + "&" \
                    "jsoncallback=_CallbackSearch&s_status=" + str(j) + "&s_tid=7" + str(m) + "&s_puin=&s_fid=&s_key=&p=" + str(i)
                r = requests.get(urls, headers=head)
                if r.status_code == 200:
                    jsons = json.loads(r.text.strip('_CallbackSearch(').strip(')'))
                for i in range(0, len(jsons["plist"])):
                    print(str(i))
                    gong = ''
                    fa = ''
                    mu = ''
                    jsonw = jsons["plist"][i]["id"]
                    num = jsonw[len(jsonw) - 2] + jsonw[len(jsonw) - 1]
                    ur = "https://scdn.gongyi.qq.com/json_data/data_detail/" + num + "/detail." + str(
                        jsonw) + ".js?jsoncallback=_cb_fn_proj_" + str(jsonw)
                    res = requests.get(ur,headers=head)
                    if res.status_code == 200:
                        nei = res.text.strip('_cb_fn_proj_').strip(str(jsonw)).strip('(').strip(');')
                        ne = json.loads(nei)
                    else:
                        num = jsonw[len(jsonw) - 1]
                        ur = "https://scdn.gongyi.qq.com/json_data/data_detail/" + num + "/detail." + str(
                            jsonw) + ".js?jsoncallback=_cb_fn_p " + str(jsonw)
                        rs = requests.get(ur,headers=head)
                        if rs.status_code == 200:
                            nei = rs.text.strip('_cb_fn_proj_').strip(str(jsonw)).strip('(').strip(');')
                            ne = json.loads(nei)
                        else:
                            num = jsonw[len(jsonw) - 2] + jsonw[len(jsonw) - 1]
                            ur = "https://scdn.gongyi.qq.com/json_data/sub_data_detail/" + num + "/detail." +str(jsonw)+ ".json"
                            rr = requests.get(ur,headers=head)
                            if rr.status_code == 200:
                                nes = json.loads(rr.text)
                                ne = nes["msg"]
                    title = ne["base"]["title"]
                    status = ne["base"]["status"]
                    startTime = ne["base"]["startTime"]
                    endTime = ne["base"]["endTime"]
                    cateName = ne["base"]["cateName"]
                    if ne["base"]["ftype"] == '1':
                        gong = ne["base"]["pName"]
                    else:
                        fa = ne["base"]["pName"]
                        mu = ne["base"]["fundName"]
                    obtainMoney = jsons["plist"][i]["stat"]["money"]
                    needMoney = ne["base"]["donate"]["needMoney"]
                    donateNum = jsons["plist"][i]["stat"]["times"]
                    desc = ne["detail"]["desc"]
                    desc = tosting(desc)
                    desc = ILLEGAL_CHARACTERS_RE.sub(r'', desc)
                    if "desc_module" in ne["detail"]:
                        if "proj_budget" in ne["detail"]["desc_module"]:
                            proj_budget = ne["detail"]["desc_module"]["proj_budget"]
                        else:
                            proj_budget = ''
                        if "proj_exe_plan" in ne["detail"]["desc_module"]:
                            proj_exe_plan = ne["detail"]["desc_module"]["proj_exe_plan"]
                        else:
                            proj_exe_plan = ''
                        if "proj_exe_content" in ne["detail"]["desc_module"]:
                            proj_exe_content = ne["detail"]["desc_module"]["proj_exe_content"]
                        else:
                            proj_exe_content = ''
                        if "proj_team_info" in ne["detail"]["desc_module"]:
                            proj_team_info = ne["detail"]["desc_module"]["proj_team_info"]
                        else:
                            proj_team_info = ''
                        if "proj_implement_res" in ne["detail"]["desc_module"]:
                            proj_implement_res = ne["detail"]["desc_module"]["proj_implement_res"]
                        else:
                            proj_implement_res = ''
                        if proj_budget is None :
                            proj_budget = ''
                        if proj_exe_plan is None:
                            proj_exe_plan = ''
                        if proj_exe_content is None:
                            proj_exe_content = ''
                        if proj_team_info is None:
                            proj_team_info = ''
                        if proj_implement_res is None:
                            proj_implement_res = ''
                    else:
                        proj_budget = ''
                        proj_exe_plan = ''
                        proj_exe_content = ''
                        proj_team_info = ''
                        proj_implement_res = ''
                    if float(needMoney) != 0.0:
                        bi = (float(obtainMoney) / 100) / (float(needMoney) / 100)
                    else:
                        bi = 0
                    if status == '1':
                        status = '募捐中'
                    elif status == '2':
                        status = '执行中'
                    else:
                        status = '已结束'
                    row = sheet.max_row
                    row = row + 1
                    sheet.cell(row, 1, title)
                    sheet.cell(row, 2, cateName)
                    sheet.cell(row, 3, status)
                    sheet.cell(row, 4, startTime + '至' + endTime)
                    sheet.cell(row, 5, startTime)
                    sheet.cell(row, 6, donateNum)
                    sheet.cell(row, 7, float(obtainMoney) / 100)
                    sheet.cell(row, 8, float(needMoney) / 100)
                    sheet.cell(row, 9, bi)
                    sheet.cell(row, 10, gong)
                    sheet.cell(row, 11, fa)
                    sheet.cell(row, 12, fa)
                    sheet.cell(row, 13, mu)
                    sheet.cell(row, 14, desc)
                    sheet.cell(row, 15, tosting(proj_budget))
                    sheet.cell(row, 16, tosting(proj_exe_plan))
                    sheet.cell(row, 17, tosting(proj_implement_res))
                    sheet.cell(row, 18, tosting(proj_exe_content))
                    sheet.cell(row, 19, tosting(proj_team_info))
                wtbook.save(excelpath)


def tosting(str):
    if str is not None:
        str = str.replace("&nbsp;", " ")
    if re.search('<img[^>]+>', str, re.I):
        src = re.findall('<img[^>]+>', str,re.I)
        for i in range(0, len(src)):
            if re.search('src="(.*?)"', src[i]):
                srcs = re.findall('src="(.*?)"', src[i], re.I)
                try:
                    res = requests.get(srcs[0])
                    if res.status_code == 200:
                        ur = srcs[0] + ' '
                except:
                    ur = 'http:' + srcs[0] + ' '
            str = str.replace(src[i],ur)
    dr = re.compile(r'<[^>]+>', re.S)
    dd = dr.sub('', str)
    return dd


if __name__ == '__main__':
    pri()
