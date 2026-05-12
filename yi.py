from selenium import webdriver
import time
from openpyxl import Workbook
import openpyxl
import requests

def VerificationClick():
    driver = webdriver.Chrome(executable_path='C:\\Users\\liu\\Desktop\\chromedriver.exe')
    # 打开首页
    url = 'https://gongyi.qq.com/succor/project_list.htm'
    driver.get(url)
    elements = driver.find_elements_by_class_name('pro_li')
    lists = set()
    for e in elements:
        name = e.find_element_by_css_selector("div[class='pro_li_img']").find_element_by_css_selector(
            "a").get_attribute("href")
        time.sleep(2)
        lists.add(name)
    for i in range(2,101):
        for e in elements:
            name = e.find_element_by_css_selector("div[class='pro_li_img']").find_element_by_css_selector(
                "a").get_attribute("href")
            time.sleep(2)
            lists.add(name)
    return lists


def excel():
    wtbook = Workbook()
    # 新增一个sheet工作表
    sheet = wtbook.create_sheet('sheet1')
    # 写入数据头
    row = 1
    col = 1
    headlist = [u'项目名称', u'项目类型（综合/济困救灾/疾病救助)', u'项目状态（募款中/执行中/已结束)', u'项目实施时间段',u'项目开始时间', u'捐款人次', u'已筹金额', u'目标金额',
                u'筹款百分比', u'项目介绍', u'项目预算', u'执行计划', u'项目效果', u'执行能力说明', u'关于我们']
    for head in headlist:
        sheet.cell(row,col,head)
        col = col + 1
    wtbook.save(r'E:\he.xlsx')


def fef(list):
    excelpath = r'E:\he.xlsx'
    wtbook = openpyxl.load_workbook(excelpath)
    sheet = wtbook.worksheets[1]
    driver = webdriver.Chrome(executable_path='C:\\Users\\liu\\Desktop\\chromedriver.exe')
    sting = ''
    for ur in list:
        driver.get(ur)
        pri = driver.find_element_by_id('money_already').text
        yi = driver.find_element_by_id('target_span').text
        name = driver.find_element_by_id('pj_name').text
        zhuang = driver.find_element_by_class_name('not_completed.current').text
        shi = driver.find_element_by_class_name('main_top_detail_target_time_value').text
        ren = driver.find_element_by_id('project_donateNum').text
        bai = driver.find_element_by_class_name('flag').text
        print("项目名称：%s,项目状态：%s,项目实施时间段：%s,捐款人次：%s,目标金额:  %s  元, 已筹金额:  %s 元，筹款百分百：%s" % (name, zhuang, shi, ren, pri, yi, bai))
        row = sheet.max_row
        row = row + 1
        sheet.cell(row, 1, name)
        sheet.cell(row, 2, '疾病救助')
        sheet.cell(row, 3, zhuang)
        sheet.cell(row, 4, shi)
        sheet.cell(row, 5, ren)
        sheet.cell(row, 6, yi)
        sheet.cell(row, 7, pri)
        sheet.cell(row, 8, bai)
        content = driver.find_element_by_xpath('//*[@id="pj_content"]')
        child = content.find_elements_by_xpath('.//*')
        i = 0
        j = 0
        while j < 6:
            cd = child.__getitem__(i)
            while cd.tag_name == 'h3' and cd.get_attribute("class") == 'title' and j < 6:
                if cd.text != '捐赠回馈':
                    i = i + 1
                    cds = child.__getitem__(i)
                    while cds.get_attribute("class") != 'title' and i < len(child):
                        if cds.tag_name == 'h3':
                            if len(cds.find_elements_by_xpath('.//*')) != 0:
                                cdd = cds.find_elements_by_xpath('.//*')
                                for cde in cdd:
                                    sting = sting + cde.text + ' '
                            else:
                                sting = sting + cds.text + ' '
                        elif cds.tag_name == 'img':
                            ur = cds.get_attribute('src')
                            sting = sting + ur + ' '
                        elif cds.tag_name == 'p':
                            sting = sting + cds.text + ' '
                        i = i + 1
                        if i < len(child) - 1:
                            cds = child.__getitem__(i)
                        else:
                            break;
                    if sting != '':
                        sheet.cell(row, j + 9, sting)
                    else:
                        if cd.tag_name == 'h3' and cd.get_attribute("class") == 'title' and j == 3:
                            sheet.cell(row, 12, sting)
                        if cd.tag_name == 'h3' and cd.get_attribute("class") == 'title' and j == 4:
                            sheet.cell(row, 13, sting)
                    cd = child.__getitem__(i)
                    sting = ''
                    j = j + 1
                else:
                    j = j + 1
    wtbook.save(excelpath)


if __name__ == '__main__':
    list = VerificationClick()
    fef(list)



