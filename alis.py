from selenium import webdriver
import time
from openpyxl import Workbook
import openpyxl
import json
import requests
from random import randint
from bs4 import BeautifulSoup
import bs4
import re




def fes():
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    excelpath = r'E:\ali.xlsx'
    wtbook = openpyxl.load_workbook(excelpath)
    sheet = wtbook.worksheets[1]
    USER_AGENTS = [
        "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; AcooBrowser; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
        "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; Acoo Browser; SLCC1; .NET CLR 2.0.50727; Media Center PC 5.0; .NET CLR 3.0.04506)",
        "Mozilla/4.0 (compatible; MSIE 7.0; AOL 9.5; AOLBuild 4337.35; Windows NT 5.1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
        "Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)",
        "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
        "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
        "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
        "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
        "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5",
        "Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.8) Gecko Fedora/1.9.0.8-1.fc10 Kazehakase/0.5.6",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
        "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52",
    ]
    random_agent = USER_AGENTS[randint(0, len(USER_AGENTS) - 1)]
    head = {
        'User-Agent': random_agent,
    }
    for i in range(1,5):
        for j in range(1,45):
            ur = "https://love.alipay.com/donate/itemList.htm?page=" + str(j) + "&&donateType=&itemClassified=" + str(i)\
                 + "&orderType=gmt_create_desc&donateShowName="
            res = requests.get(ur)
            soup = BeautifulSoup(res.text, 'html.parser')
            lis = soup.select("li[class='donate-item-default-li fn-clear']")
            for li in lis:
                type = li.select("span[class='ft-green']")
                donateType = type[0].text
                rens = type[1].text
                ur = li.find(name='a',attrs={'class:','donate-item-default-more'}).get('href')
                ren = li.find(name='dd',attrs={'class:','donate-item-default-dd'}).text
                shi = li.find(name='p',attrs={'class:','donate-item-default-dd'})
                shis = shi.string.split(sep='至')[0].split(sep='：')[1]
                yi = li.find(name='em',attrs={'class:','ft-orange ft-bold'}).text
                urs = ur.split(sep='?')
                name = urs[1]
                sting = ''
                url = "https://love.alipay.com/donate/showFeedBack.json?" + name + "&page=1"
                responses = requests.get(url, headers={'referer': ur})
                jse = json.loads(responses.text)
                if jse['stat'] == 'ok':
                    for m in range(1,30):
                        url = "https://love.alipay.com/donate/showFeedBack.json?" + name + "&page=" + str(m)
                        response = requests.get(url, headers={'referer': ur})
                        js = json.loads(response.text)
                        if js['stat'] == 'ok':
                            tt = js['donateItemModel']['donateTitle']
                            nei = js['donateItemModel']['editorIntro']
                            bai = js['donateItemModel']['publicFundraisingNo']
                            list = js['donateFeedbackPageModelList']
                            for n in range(0,len(list)):
                                lis = list[n]
                                moder = lis['donateFeedbackProgressModel']['gmtAuditStr']
                                title = lis['donateFeedbackProgressModel']['title']
                                xi = lis['donateFeedbackProgressModel']['feedback']
                                sting = sting + tosting(moder) + ' '
                                sting = sting + tosting(title) + ' '
                                sting = sting + tosting(xi) + ' '
                                if 'bigPicPathList' in lis:
                                    bics = lis['bigPicPathList']
                                    for x in range(0,len(bics)):
                                        ur = bics[x]['picViewUrl']
                                        sting = sting + ur + ' '
                    sting = ILLEGAL_CHARACTERS_RE.sub(r'', sting)
                    row = sheet.max_row
                    row = row + 1
                    sheet.cell(row, 1, donateType)
                    sheet.cell(row, 2, tt)
                    sheet.cell(row, 3, yi)
                    sheet.cell(row, 4, bai)
                    sheet.cell(row, 5, rens)
                    sheet.cell(row, 6, shi.text)
                    sheet.cell(row, 7, shis)
                    sheet.cell(row, 8, ren)
                    sheet.cell(row, 9, tosting(nei))
                    sheet.cell(row, 10,sting)
            wtbook.save(excelpath)


def tosting(str):
    if str is not None:
        str = str.replace("&nbsp;", " ")
    if re.search('<img[^>]+>', str, re.I):
        src = re.findall('<img[^>]+>', str,re.I)
        for i in range(0, len(src)):
            if re.search('src="(.*?)"', src[i]):
                srcs = re.findall('src="(.*?)"', src[i], re.I)
                try:
                    res = requests.get(srcs[0])
                    if res.status_code == 200:
                        ur = srcs[0] + ' '
                except:
                    ur = 'http:' + srcs[0] + ' '
            str = str.replace(src[i],ur)
    dr = re.compile(r'<[^>]+>', re.S)
    dd = dr.sub('', str)
    return dd


def excel():
    wtbook = Workbook()
    # 新增一个sheet工作表
    sheet = wtbook.create_sheet('sheet1')
    # 写入数据头
    row = 1
    col = 1
    headlist = [u'项目类型（教育助学/扶困救灾/医疗救助/其他)', u'项目名称', u'已筹善款', u'募捐方案备案编号',
                u'参捐人数', u'项目时间', u'开始时间', u'发布机构', u'项目介绍', u'善款去向']
    for head in headlist:
        sheet.cell(row,col,head)
        col = col + 1
    wtbook.save(r'E:\ali.xlsx')


if __name__ == '__main__':
    fes()