import openpyxl
from selenium import webdriver
import xlwt
class fe():
    driver = webdriver.Chrome(executable_path='C:\\Users\\liu\\Desktop\\chromedriver.exe')
    # 打开首页
    url = 'https://gongyi.weibo.com/r/241924'
    driver.get(url)
    name = driver.find_element_by_class_name('tit').find_element_by_css_selector('strong').text
    pri = driver.find_element_by_class_name('num-left').find_element_by_class_name('m-fontWeight').text
    yi = driver.find_element_by_class_name('num-center').find_element_by_class_name('m-fontWeight').text
    print("项目名称： %s,已筹:  %s  元, 捐款人次:  %s 次" % (name, pri, yi))
    xiang = driver.find_element_by_id('help_info')
    child = xiang.find_elements_by_xpath('.//*')
    sting = ''
    faq = ''
    mu = ''
    jin = ''
    for id in child:
        if id.tag_name == 'div':
            if len(id.find_elements_by_xpath('.//*')) != 0:
                continue;
            else:
                sting = sting + id.text + ' '
        if id.tag_name == 'h2':
            sting = sting + id.text + ' '
        if id.tag_name == 'b':
            sting = sting + id.text + ' '
        if id.tag_name == 'img':
            sting = sting + id.get_attribute('src') + ' '
    print(sting)
    fa = driver.find_element_by_class_name('tab_b_con')
    fas = fa.find_element_by_css_selector('h1')
    faq = faq + fas.text
    fe = driver.find_element_by_class_name('b_con_conL')
    fes = driver.find_element_by_class_name('b_con_conR')
    fed = fe.find_elements_by_xpath('.//*')
    for fd in fed:
        if fd.tag_name == 'h2':
            faq = faq + fd.text + ' '
        else:
            ur = fd.find_element_by_css_selector('a').find_element_by_css_selector('img').get_attribute('src')
            faq = faq + ur + ' '
            faq = faq + fd.text + ' '
            break;
    fesd = fes.find_elements_by_xpath('.//*')
    for fds in fesd:
        if fds.tag_name == 'h2':
            faq = faq + fds.text + ' '
        else:
            ur = fds.find_element_by_css_selector('a').find_element_by_css_selector('img').get_attribute('src')
            faq = faq + ur + ' '
            h5 = fds.find_element_by_css_selector('h5')
            h6 = fds.find_element_by_css_selector('h6')
            faq = faq + h5.text + ' '
            faq = faq + h6.text + ' '
            p = fds.find_element_by_css_selector('p')
            ps = p.find_elements_by_xpath('.//*')
            for psd in ps:
                if psd.tag_name == 'span':
                    faq = faq + psd.text + ' '
                if psd.tag_name == 'strong':
                    if len(psd.find_elements_by_xpath('.//*')) != 0:
                        a = psd.find_element_by_css_selector('a')
                        faq = faq + a.get_attribute('url_open') + ' '
                    else:
                        faq = faq + psd.text + ' '
            break;
    print(faq)
    a = driver.find_element_by_id('loading_donate')
    dl = driver.find_elements_by_class_name('dynamic_con.clearfix')
    for dll in dl:
        ur = dll.find_element_by_class_name('m-img-box').find_element_by_css_selector('img')
        p1 = dll.find_element_by_class_name('dynamic_name')
        p2 = dll.find_element_by_class_name('dynamic_time')
        p3 = dll.find_element_by_class_name('dynamic_txt')
        p4 = dll.find_element_by_class_name('dynamic_num')
        mu = mu + ur.get_attribute('src') + ' ' + p1.text + ' ' + p2.text + ' ' + p3.text + ' ' + p4.text + ' '
    print(mu)
    je = driver.find_element_by_class_name('line_time').find_element_by_css_selector('span') \
        .find_element_by_css_selector('strong')
    jes = driver.find_element_by_class_name('line_time').find_element_by_css_selector('span') \
        .find_element_by_css_selector('em')
    jd = driver.find_element_by_class_name('txt')
    jds = driver.find_element_by_class_name('add').find_element_by_css_selector('em')
    jin = jin + je.text + ' ' + jes.text + ' ' + jd.text + ' ' + jds.text + ' '
    print(jin)
    excelpath = r'E:\hl.xlsx'
    wtbook = openpyxl.load_workbook(excelpath)
    sheet = wtbook.worksheets[1]
    row = sheet.max_row
    row = row + 1
    sheet.cell(row, 1, name)
    sheet.cell(row, 2, pri)
    sheet.cell(row, 3, yi)
    sheet.cell(row, 4, sting)
    sheet.cell(row, 5, faq)
    sheet.cell(row, 6, mu)
    sheet.cell(row, 7, jin)
    wtbook.save(excelpath)



if __name__ == '__main__':
    fe()


